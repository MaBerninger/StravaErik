#!/usr/bin/env python3
"""
export_strava.py
Gebruik: python export_strava.py [pad_naar_excel] [output_json]

Standaard:
  excel:  strava_analyse.xlsx
  output: strava_data.json

Zet dit script in dezelfde map als je Excel bestand.
Run het elke keer als je Excel is bijgewerkt, of automatiseer via Task Scheduler / cron.
"""

import sys
import json
import numpy as np
import pandas as pd
import openpyxl
from pathlib import Path

EXCEL_PATH = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("strava_analyse.xlsx")
OUTPUT_PATH = Path(sys.argv[2]) if len(sys.argv) > 2 else Path("strava_data.json")

def tempo_to_decimal(t):
    try:
        parts = str(t).split(":")
        return int(parts[0]) + int(parts[1]) / 60
    except:
        return None

def min_to_time(m):
    if not m or (isinstance(m, float) and np.isnan(m)) or m == 0:
        return "—"
    h, rem = divmod(int(m), 60)
    s = int((m % 1) * 60)
    return f"{h}:{rem:02d}:{s:02d}" if h else f"{rem}:{s:02d}"

def clean(v):
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return ""
    return str(v).strip()

def nan_safe_int(v):
    try:
        f = float(v)
        return int(f) if not np.isnan(f) else 0
    except:
        return 0

print(f"[>>] Excel inladen: {EXCEL_PATH}")

# ── 1. Hyperlinks via openpyxl ──────────────────────────────────────────────
wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb["Trainingsdata"]
header_row = 11
cols = {cell.value: cell.column - 1 for cell in ws[header_row] if cell.value}

link_map = {}
for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
    seg = str(row[cols.get("Segment", 4)].value or "")
    if "Totale Run" not in seg:
        continue
    datum = str(row[cols.get("Datum", 1)].value or "").strip()
    naam = str(row[cols.get("Run Naam", 2)].value or "").strip()
    link_cell = row[cols.get("Link", 3)]
    url = link_cell.hyperlink.target if link_cell.hyperlink else ""
    key = f"{datum}|{naam}"
    if key not in link_map:
        link_map[key] = []
    if url:
        link_map[key].append(url)

# ── 2. Data via pandas ───────────────────────────────────────────────────────
df = pd.read_excel(EXCEL_PATH, sheet_name="Trainingsdata", header=10)
df["Afstand (km)"] = pd.to_numeric(df["Afstand (km)"], errors="coerce")
df["Gem HR"] = pd.to_numeric(df["Gem HR"], errors="coerce")
df["Min HR"] = pd.to_numeric(df["Min HR"], errors="coerce")
df["Max HR"] = pd.to_numeric(df["Max HR"], errors="coerce")
df["Run Naam"] = df["Run Naam"].ffill()
df["Datum"] = df["Datum"].ffill()
df["Is_totaal"] = df["Segment"].str.strip() == "🏃\u200d♂️ Totale Run"
df["Run_afstand"] = np.where(df["Is_totaal"], df["Afstand (km)"], np.nan)
df["Run_afstand"] = df["Run_afstand"].ffill()

# ── 3. Rondes per run (voor HR@5:00 berekening) ──────────────────────────────
rondes = df[df["Segment"].str.strip().str.startswith("Ronde", na=False)].copy()
rondes["Ronde_nr"] = rondes["Segment"].str.extract(r"(\d+)").astype(float)
rondes["Tempo_dec"] = rondes["Tempo (min/km)"].apply(tempo_to_decimal)

SLOPE = -9.71
DRIFT = 0.29

hr500_map = {}
for (datum, naam), group in rondes.groupby(["Datum", "Run Naam"]):
    run_afstand = group["Run_afstand"].iloc[0]
    if run_afstand <= 8:
        continue
    g_all = group[group["Ronde_nr"] >= 3].copy()
    if any(g_all["Afstand (km)"].fillna(1) < 0.4) and (g_all["Afstand (km)"] < 0.4).sum() >= 3:
        continue
    g = g_all[
        (g_all["Tempo_dec"] >= 4.0) &
        (g_all["Tempo_dec"] <= 6.0) &
        g_all["Gem HR"].notna()
    ]
    if len(g) < 3:
        continue
    gem_tempo = g["Tempo_dec"].mean()
    gem_hr = g["Gem HR"].mean()
    midpoint_km = (3 + run_afstand) / 2
    drift_corr = -(midpoint_km - 3) * DRIFT
    hr_500 = gem_hr + drift_corr + SLOPE * (5.0 - gem_tempo)
    if 100 < hr_500 < 185:
        hr500_map[f"{datum}|{naam}"] = round(hr_500, 1)

# ── 4. Alle runs bouwen ───────────────────────────────────────────────────────
totaal = df[df["Is_totaal"] & (df["Afstand (km)"] > 0.5)].copy()
totaal["Tempo_dec"] = totaal["Tempo (min/km)"].apply(tempo_to_decimal)
totaal["Tijd_min"] = totaal["Tempo_dec"] * totaal["Afstand (km)"]
totaal["Tijd_str"] = totaal["Tijd_min"].apply(min_to_time)

alle_runs = []
for _, r in totaal.sort_values("Datum", ascending=False).iterrows():
    datum = clean(r["Datum"])
    naam = clean(r["Run Naam"])
    key = f"{datum}|{naam}"
    urls = link_map.get(key, [])
    url = urls.pop(0) if urls else ""
    link_map[key] = urls
    afstand = float(r["Afstand (km)"]) if not np.isnan(r["Afstand (km)"]) else 0
    alle_runs.append({
        "datum": datum,
        "datum_sort": datum.split(",")[0].strip(),
        "naam": naam,
        "afstand": round(afstand, 2),
        "tempo": clean(r["Tempo (min/km)"]),
        "tempo_dec": round(r["Tempo_dec"], 3) if r["Tempo_dec"] else 0,
        "tijd": r["Tijd_str"],
        "min_hr": nan_safe_int(r["Min HR"]),
        "gem_hr": nan_safe_int(r["Gem HR"]),
        "max_hr": nan_safe_int(r["Max HR"]),
        "zone": clean(r["HR Zone"]),
        "trainingsplan": clean(r["Trainingsplan"]),
        "strava_url": url,
        "hr_500": hr500_map.get(key, None),
    })

# ── 5. Duurlopen & wedstrijden categoriseren ─────────────────────────────────
def is_wedstrijd(r):
    naam = r["naam"].lower()
    plan = r["trainingsplan"].lower()
    kw = ["marathon rotterdam", "haas", "silvester", "schoorl", "cpc 21", "vaartspel spanderswoud"]
    return any(k in naam or k in plan for k in kw)

def is_duurloop(r):
    return (not is_wedstrijd(r) and r["afstand"] >= 8 and
            ("Z1" in r["zone"] or "Z2" in r["zone"] or "duurloop" in r["trainingsplan"].lower()))

duurlopen = [r for r in alle_runs if is_duurloop(r)]
hr500_list = [{"datum": r["datum_sort"], "hr500": r["hr_500"], "afstand": r["afstand"],
               "gem_hr": r["gem_hr"], "gem_tempo": r["tempo_dec"]}
              for r in duurlopen if r["hr_500"]]

# ── 6. Wedstrijd events — dynamisch detecteren uit naam + trainingsplan ────────
import re as _re
from itertools import groupby as _groupby

# Wedstrijdkeywords → event naam
WEDSTRIJD_KEYWORDS = [
    (["marathon rotterdam", "marathondag"],          "Marathon Rotterdam"),
    (["haas ultramarathon", "haas voor"],            "Haas"),
    (["haas"],                                        "Haas"),
    (["silvestercross"],                              "Silvestercross"),
    (["schoorl"],                                     "Schoorl 10 km"),
    (["cpc 21", "cpc 10"],                            "CPC"),
    (["vaartspel spanderswoud"],                      "Vaartspel Spanderswoud"),
    (["vaartspel bosberg"],                           "Vaartspel Bosberg"),
    (["zomeravondcup"],                               "Zomeravondcup"),
    (["de mooiste", "#de mooiste"],                   "De Mooiste"),
]

def detecteer_event_naam(r):
    tekst = (r["naam"] + " " + r["trainingsplan"]).lower()
    for kws, naam in WEDSTRIJD_KEYWORDS:
        if any(k in tekst for k in kws):
            return naam
    return None

# Groepeer wedstrijdruns per event+datum
event_map = {}
for r in alle_runs:
    if r["gem_hr"] < 155 and r["afstand"] < 5:
        continue  # te rustig / te kort voor wedstrijd
    naam = detecteer_event_naam(r)
    if naam:
        # Gebruik naam + week als key zodat verschillende edities apart blijven
        week = r["datum_sort"][:7]  # YYYY-MM
        key = f"{naam}|{week}"
        if key not in event_map:
            event_map[key] = {"naam": naam, "runs": [], "_datum_sort": r["datum_sort"]}
        event_map[key]["runs"].append(r)

# Bouw events lijst
events = []
for key, e in sorted(event_map.items(), key=lambda x: x[1]["_datum_sort"], reverse=True):
    runs = e["runs"]
    # Wedstrijdrun = de run met hoogste gem HR
    race = max((r for r in runs if r["gem_hr"] > 100), key=lambda x: x["gem_hr"], default=None)
    if not race:
        continue

    # Datum in leesbare vorm
    try:
        from datetime import datetime as _dt
        d = _dt.strptime(race["datum_sort"], "%Y-%m-%d")
        maanden = ["","januari","februari","maart","april","mei","juni",
                   "juli","augustus","september","oktober","november","december"]
        datum_nl = f"{d.day} {maanden[d.month]} {d.year}"
    except:
        datum_nl = race["datum_sort"]

    events.append({
        "naam": e["naam"],
        "datum": datum_nl,
        "afstand": round(race["afstand"], 2),
        "gem_hr": race["gem_hr"],
        "max_hr": race["max_hr"],
        "tempo": race["tempo"],
        "tempo_dec": race["tempo_dec"],
        "tijd": race["tijd"],
        "strava_url": race["strava_url"],
        "runs": runs,
    })

# ── 7. PR's — dynamisch uit Strava bio parsen ──────────────────────────────────
def parse_prs_uit_bio(bio):
    """Parst PRs uit bio zoals: PRs 1,5km 4:22, 3km 8:59, 5km 15:30 ..."""
    prs = []
    if not bio:
        return prs
    # Zoek patronen: afstand + tijd
    pattern = r'([\d,\.]+\s*km)\s+([\d]+:[\d]{2}(?::[\d]{2})?)'
    matches = _re.findall(pattern, bio, _re.IGNORECASE)
    for afstand_str, tijd_str in matches:
        afstand_str = afstand_str.strip()
        # Bereken tempo
        try:
            afstand_km = float(afstand_str.replace('km','').replace(',','.').strip())
            parts = tijd_str.split(':')
            if len(parts) == 3:
                total_sec = int(parts[0])*3600 + int(parts[1])*60 + int(parts[2])
            else:
                total_sec = int(parts[0])*60 + int(parts[1])
            tempo_sec = total_sec / afstand_km
            tempo_str = f"{int(tempo_sec//60)}:{int(tempo_sec%60):02d}"
        except:
            tempo_str = ""
        prs.append({"afstand": afstand_str, "tijd": tijd_str, "tempo": tempo_str, "datum": ""})
    return prs

# Lees bio uit Excel
try:
    import openpyxl as _opx
    _wb = _opx.load_workbook(EXCEL_PATH, data_only=True)
    _ws = _wb[_wb.sheetnames[0]]
    bio_text = str(_ws['A2'].value or '')
    prs = parse_prs_uit_bio(bio_text)
    print(f"[OK] {len(prs)} PRs geparsed uit bio")
except Exception as _e:
    print(f"[!] Bio lezen mislukt: {_e}, gebruik fallback PRs")
    prs = []

# Fallback als bio leeg is
if not prs:
    prs = [
        {"afstand": "1,5 km", "tijd": "4:22", "tempo": "2:55", "datum": ""},
        {"afstand": "5 km",   "tijd": "16:53", "tempo": "3:23", "datum": ""},
        {"afstand": "10 km",  "tijd": "32:12", "tempo": "3:13", "datum": ""},
        {"afstand": "21,1 km","tijd": "1:11:15","tempo": "3:22","datum": ""},
        {"afstand": "42,2 km","tijd": "2:36:07","tempo": "3:41","datum": "Rotterdam 2026"},
    ]

# ── 8. Exporteer JSON ─────────────────────────────────────────────────────────
output = {
    "gegenereerd_op": pd.Timestamp.now().strftime("%Y-%m-%d %H:%M"),
    "alle": alle_runs,
    "duurlopen": duurlopen,
    "hr500": hr500_list,
    "events": events,
    "prs": prs,
}

with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
    json.dump(output, f, ensure_ascii=False, indent=2)

print(f"[OK] Klaar!")
print(f"   Runs:       {len(alle_runs)}")
print(f"   Duurlopen:  {len(duurlopen)}")
print(f"   HR@5:00:    {len(hr500_list)}")
print(f"   Output:     {OUTPUT_PATH}")